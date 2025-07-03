# File: src/financial_analysis/reporting/excel_reporter.py
# Purpose: Generates a professional, multi-sheet, and formatted Excel report.

import logging
import pandas as pd
from .base_reporter import BaseReporter
from ..core.models import CompanyAnalysis

# Import openpyxl types for styling
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

class ExcelReporter(BaseReporter):
    """Generates a professional, formatted report in an .xlsx file."""

    def generate_report(self, analysis: CompanyAnalysis, output_path: str) -> None:
        logger.info(f"Generating Excel report for {analysis.company_info.ticker} to {output_path}")
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                self._write_summary_sheet(writer, analysis)
                self._write_ratios_sheet(writer, analysis)
                self._write_statements_sheet(writer, analysis)
            logger.info("Successfully generated Excel report.")
        except IOError as e:
            logger.error(f"Failed to write Excel report to {output_path}: {e}")

    def _apply_styles(self, ws: Worksheet, df: pd.DataFrame):
        """Applies formatting to a worksheet."""
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
        
        # Style headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Auto-fit column widths
        for i, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            ws.column_dimensions[chr(65 + i)].width = max_len
            
        # Freeze top row
        ws.freeze_panes = 'A2'

        # Apply number formats based on column name
        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = chr(64 + col_idx)
            if 'ratio' in col_name.lower() or 'turnover' in col_name.lower() or 'equity' in col_name.lower():
                num_format = '0.00'
            elif 'margin' in col_name.lower() or 'roe' in col_name.lower() or 'roa' in col_name.lower():
                num_format = '0.00%'
            elif any(k in col_name.lower() for k in ['revenue', 'profit', 'income', 'assets', 'liabilities', 'cash', 'debt', 'equity']):
                num_format = '"$"#,##0'
            else:
                continue
            
            for row_idx in range(2, ws.max_row + 1):
                ws[f'{col_letter}{row_idx}'].number_format = num_format
    
    def _write_summary_sheet(self, writer: pd.ExcelWriter, analysis: CompanyAnalysis):
        """Writes the summary and qualitative analysis sheet."""
        info = analysis.company_info
        qual = analysis.qualitative_analysis
        
        summary_data = {
            "Metric": [
                "Company", "Ticker", "Sector", "Industry", "Analysis Date", "",
                "Overall Summary", "Key Strengths", "Key Areas for Attention", "",
                "Liquidity Analysis", "Profitability Analysis", "Leverage Analysis", "Efficiency Analysis", "",
                "Disclaimer"
            ],
            "Value": [
                info.name, info.ticker, info.sector, info.industry, analysis.analysis_date.strftime('%Y-%m-%d'), "",
                qual.get('overall_summary', 'N/A'), "\n".join(qual.get('key_strengths', [])), "\n".join(qual.get('key_concerns', [])), "",
                qual.get('liquidity', 'N/A'), qual.get('profitability', 'N/A'), qual.get('leverage', 'N/A'), qual.get('efficiency', 'N/A'), "",
                "For educational purposes only. Not financial advice."
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="Summary & Analysis", index=False)
        
        # Apply styles
        ws = writer.sheets["Summary & Analysis"]
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 100
        for row in ws.iter_rows(min_row=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    def _write_ratios_sheet(self, writer: pd.ExcelWriter, analysis: CompanyAnalysis):
        """Writes the historical financial ratios sheet."""
        if not analysis.historical_ratios:
            return
            
        df_ratios = pd.DataFrame([r.model_dump() for r in analysis.historical_ratios])
        df_ratios = df_ratios.set_index(['fiscal_year', 'period']).drop(columns=['ticker'])
        df_ratios.to_excel(writer, sheet_name="Financial Ratios")
        self._apply_styles(writer.sheets["Financial Ratios"], df_ratios.reset_index())

    def _write_statements_sheet(self, writer: pd.ExcelWriter, analysis: CompanyAnalysis):
        """Writes the historical financial statements sheet."""
        if not analysis.historical_statements:
            return
            
        flat_data = []
        for stmt in analysis.historical_statements:
            row = {"fiscal_year": stmt.fiscal_year, "end_date": stmt.end_date.strftime('%Y-%m-%d')}
            row.update(stmt.income_statement.model_dump())
            row.update(stmt.balance_sheet.model_dump())
            row.update(stmt.cash_flow_statement.model_dump())
            flat_data.append(row)
            
        df_statements = pd.DataFrame(flat_data)
        df_statements = df_statements.set_index('fiscal_year')
        df_statements.to_excel(writer, sheet_name="Financial Statements")
        self._apply_styles(writer.sheets["Financial Statements"], df_statements.reset_index())
